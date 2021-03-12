from io import StringIO
from io import BytesIO

import pandas as pd
#from xlwt import Workbook, easyxf

from pydap.model import *
from pydap.lib import walk
from pydap.responses.lib import BaseResponse
import tempfile
from functools import singledispatch
from openpyxl import Workbook

class ExcelResponse(BaseResponse):

    __description__ = "Excel spreadsheet"

    def __init__(self, dataset):
        BaseResponse.__init__(self, dataset)
        self.headers.extend([
                ('Content-description', 'dods_xls'),
                ('Content-type', 'application/vnd.ms-excel'),
                ])
    
    def __iter__(self):
        dataset = self.dataset
        #buf = StringIO()
        buf = BytesIO()
        wb = Workbook()
        #ws = wb.create_sheet('Global attributes')
        #write_metadata(ws, self.dataset, 0, 0)

        for seq in walk(dataset, SequenceType):
            ws = wb.create_sheet(seq.id)

            # add header 
#            for j, var_ in enumerate(seq.keys()):
#                ws.write(0, j, var_)

            # add data
            for i, row in enumerate(seq.data):
                #for j, value in enumerate(seq.data):
                    #for j, value in enumerate(row):
                        ws.append([value for value in row])
                        #ws.write(i+1, j, value)

            # add var metadata
            #n = 0
            #j = len(seq.keys())+1
            #for child in seq.walk():
                #ws.write_merge(n, n, j, j+1, child.name, HEADER)
#                ws.write(n, n, j, j+1, child.name, HEADER)
#                n = write_metadata(ws, child, n+1, j)+1
            #    pass
        wb.save(buf)
        print(buf)
        yield buf.getvalue()
        return [ buf.getvalue() ]


#        df = pd.DataFrame({'foo':[1,2,3,4], 'bar':[54,32,91, 57]})
#        temp_file = tempfile.NamedTempFile()
#        df.to_excel(temp_file)

        wb.save(buf)
        return [ buf.getvalue() ]

        with open(temp_file, 'rb') as f: 
            yield f.read()

    @staticmethod
    def serialize(dataset):
        buf = StringIO()
        df = pd.DataFrame({'foo':[1,2,3,4], 'bar':[54,32,91, 57]})
        df.to_xslx(buf)
        return [ buf.getvalue()]

        wb = Workbook()

        # dataset metadata
        ws = wb.add_sheet('Global attributes')
        write_metadata(ws, dataset, 0, 0)

        # sequences
        for seq in walk(dataset, SequenceType):
            ws = wb.add_sheet(seq.name)

            # add header
            for j, var_ in enumerate(seq.keys()):
                ws.write(0, j, var_, HEADER)

            # add data
            for i, row in enumerate(seq.data):
                for j, value in enumerate(row):
                    ws.write(i+1, j, value)

            # add var metadata
            n = 0
            j = len(seq.keys())+1
            for child in seq.walk():
                ws.write_merge(n, n, j, j+1, child.name, HEADER)
                n = write_metadata(ws, child, n+1, j)+1

        wb.save(buf)
        return [ buf.getvalue() ]


def write_metadata(ws, var, i, j):
    for k, v in var.attributes.items():
        n = height(v)
        write_attr(ws, k, v, i, j)
        i += n
    return i


def write_attr(ws, k, v, i, j):
    if isinstance(v, dict):
        n = height(v)
        ws.write(i, i+n-1, j, j, '%s:' % k, HEADER)
#        ws.write_merge(i, i+n-1, j, j, '%s:' % k, HEADER)
        ws.col(j).width = max(ws.col(j).width, (len(k)+1)*350)
        for kk, vv in v.items():
            n = height(vv)
            write_attr(ws, kk, vv, i, j+1)
            i += n
    else:
        ws.write(i, j, '%s:' % k, HEADER)
        ws.col(j).width = max(ws.col(j).width, (len(k)+1)*350)
        ws.write(i, j+1, str(v))
        ws.col(j+1).width = max(ws.col(j+1).width, len(str(v))*350)


def height(v):
    """Return the number of elements in an attribute."""
    if isinstance(v, dict):
        return sum( height(o) for o in v.items() )
    else:
        return 1
